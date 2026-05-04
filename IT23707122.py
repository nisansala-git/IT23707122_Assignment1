"""
Singlish → Sinhala transliteration tester - Frontend UI only
Waits for clipboard to actually change before reading it.
"""

import argparse, subprocess, time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from playwright.sync_api import sync_playwright

def mac_paste():
    try:
        r = subprocess.run(['pbpaste'], capture_output=True, text=True, encoding='utf-8')
        return r.stdout.strip()
    except:
        return ""

def mac_copy(text):
    """Put known text into clipboard so we can detect when it changes."""
    try:
        subprocess.run(['pbcopy'], input=text, text=True, encoding='utf-8')
    except:
        pass

def wait_for_new_clipboard(sentinel, timeout_sec=12):
    """Wait until clipboard changes from the sentinel value."""
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        current = mac_paste()
        if current and current != sentinel and len(current) > 0:
            return current
        time.sleep(0.5)
    return ""

def run_test():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--url", required=True)
    parser.add_argument("--wait-ms", default=8000, type=int)
    parser.add_argument("--type-delay-ms", default=80, type=int)
    parser.add_argument("--slow-mo-ms", default=200, type=int)
    parser.add_argument("--input-col", default="Input")
    parser.add_argument("--expected-col", default="Expected output")
    parser.add_argument("--actual-col", default="Actual output")
    parser.add_argument("--status-col", default="Status")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel)
    ws = wb.active

    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    input_col    = headers.get(args.input_col, 3)
    expected_col = headers.get(args.expected_col, 4)
    actual_col   = headers.get(args.actual_col, 5)
    status_col   = headers.get(args.status_col, 6)

    print(f"Starting test with {ws.max_row - 1} rows...\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            slow_mo=args.slow_mo_ms,
            args=["--disable-web-security", "--no-sandbox"]
        )
        context = browser.new_context(ignore_https_errors=True)
        page = context.new_page()

        # Log API calls for info only
        def on_response(response):
            if 'translate' in response.url or 'singlish' in response.url:
                print(f"  [API {response.status}]")
        page.on("response", on_response)

        page.goto(args.url, wait_until="networkidle")
        page.wait_for_timeout(3000)

        # Switch to Chat Sinhala via top menu
        print("Switching to Chat Sinhala mode...")
        try:
            page.get_by_role("button", name="Transliteration").click()
            page.wait_for_timeout(1000)
            page.get_by_text("Chat Sinhala").click()
            page.wait_for_timeout(2000)
            print("Chat Sinhala mode selected!\n")
        except Exception as e:
            print(f"Warning: {e}\n")

        input_ta  = page.locator('textarea').first
        output_ta = page.locator('textarea').nth(1)
        trans_btn = page.locator('button:has-text("Transliterate")').first

        for row in range(2, ws.max_row + 1):
            singlish = str(ws.cell(row, input_col).value or "").strip()
            expected = str(ws.cell(row, expected_col).value or "").strip()
            if not singlish:
                continue

            print(f"Testing [Row {row}]: {singlish[:60]}")

            try:
                # 1. Set a unique sentinel in clipboard
                sentinel = f"__SENTINEL_{row}__"
                mac_copy(sentinel)

                # 2. Clear input
                input_ta.click()
                page.keyboard.press("Meta+a")
                page.wait_for_timeout(300)
                page.keyboard.press("Backspace")
                page.wait_for_timeout(400)

                # 3. Type singlish
                input_ta.fill(singlish)
                page.wait_for_timeout(500)

                # 4. Click Transliterate
                if trans_btn.count() > 0 and trans_btn.is_enabled():
                    trans_btn.click()
                    print("  Clicked Transliterate")

                # 5. Wait for API to respond
                page.wait_for_timeout(args.wait_ms)

                # 6. Click output box, select all, copy
                output_ta.click()
                page.wait_for_timeout(600)
                page.keyboard.press("Meta+a")
                page.wait_for_timeout(400)
                page.keyboard.press("Meta+c")
                page.wait_for_timeout(600)

                # 7. Wait until clipboard actually changes from sentinel
                actual = wait_for_new_clipboard(sentinel, timeout_sec=10)
                print(f"  Got output: '{actual[:70]}'")

                # 8. Write to Excel
                cell_actual = ws.cell(row, actual_col, value=actual)
                cell_actual.alignment = Alignment(wrap_text=True, vertical="center")
                cell_actual.font = Font(name="Arial", size=9)

                status = "FAIL" if (not actual or actual != expected) else "PASS"

                cell_status = ws.cell(row, status_col, value=status)
                cell_status.font = Font(name="Arial", size=9, bold=(status == "FAIL"))
                cell_status.alignment = Alignment(horizontal="center", vertical="center")
                cell_status.fill = PatternFill("solid", fgColor="FFCCCC" if status == "FAIL" else "CCFFCC")

                print(f"  -> {status}\n")
                wb.save(args.excel)

            except Exception as e:
                print(f"  ERROR: {e}\n")
                ws.cell(row, status_col, value="UI Error")
                wb.save(args.excel)

        print("All done! Excel saved.")
        browser.close()

if __name__ == "__main__":
    run_test()
